import openpyxl
import json

#===============================================
#  Json файла с чеками

with open('C:\\Users\\Alex_job\\Documents\\Git\\Test_site\\programma_dly_obrabotki_chekov\\extract.json', mode= 'r', encoding= 'UTF-8') as file:
    my_file = json.load(file)
    new_json = json.dumps(my_file, ensure_ascii=False, indent= 4)
    my_json = json.loads(new_json)


#===============================================
#  Сделал для теста    

cheque = my_json[0]

data_purchase = cheque['createdAt']
product_name = cheque["ticket"]["document"]["receipt"]["items"][1]["name"]
price_product = cheque["ticket"]["document"]["receipt"]["items"][0]["price"]
quantity = cheque["ticket"]["document"]["receipt"]["items"][0]["quantity"]

count_items = len(cheque["ticket"]["document"]["receipt"]["items"])


#===============================================
#  Цикл создания словаря из нужных значений

dict_product = {}
list_product = []


for i in range(count_items):
    #dict_product = {cheque["ticket"]["document"]["receipt"]["items"][i]["name"] : [cheque["ticket"]["document"]["receipt"]["items"][i]["price"], cheque["ticket"]["document"]["receipt"]["items"][i]["quantity"]]}
    dict_product[cheque["ticket"]["document"]["receipt"]["items"][i]["name"]] = [cheque["ticket"]["document"]["receipt"]["items"][i]["price"] / 100,
                                                                                  cheque["ticket"]["document"]["receipt"]["items"][i]["quantity"], 
                                                                                  cheque['createdAt'][0:-15]]


#===============================================
#  Кадры для созданрия файла exel и его заполнения (шаблон)

#wb = Workbook()
#income = wb.active
#expenses = wb.create_sheet('Расходы')
#income.title = 'Доходы'
#income['A1'] = product_name
#expenses['A1'] = 'Январь'
#expenses['A2'] = 'Категория'
#expenses['B2'] = 'Подкатегория'
#expenses['A3'] = 'Продукты'
celendar = {1:'c', 2:'d', 3:'e', 4:'f', 5:'g', 6:'h', 7:'i', 8:'j', 9:'k', 10:'l', 11:'m', 12:'n',
             13:'o', 14:'p', 15:'q', 16:'r', 17:'s', 18:'t', 19:'u', 20:'v', 21:'w', 22:'x', 23:'y',
             24:'z', 25:'aa', 26:'ab', 27:'ac', 28:'ad', 29:'ae', 30:'af', 31:'ag'}
#print(wb.sheetnames)

#===============================================
#  Открытие сущетвующего файла exel и его наполнение чеками

wb = openpyxl.load_workbook(filename= 'test.xlsx')

expenses = wb['Расходы']
income = wb['Доходы']

count = 2

for product in dict_product.keys():
    month = int(dict_product[product][2][5:7])
    day = int(dict_product[product][2][8:])
    if expenses[f'B{count}'] != None:
        expenses[f'B{count}'] = product
        expenses[f'{celendar[day]}{count}'] = dict_product[product][0] #price
    i += 1


class Check:

    count = 2
    
    def __init__(self, check):
        self.dict_product = check
        self.scaning_products(self)

    def scaning_products(self, element):
        for key in self.dict_product:
            if key in list_product:
                self.add_price(self, count)
            else:
                list_product.append(key)
                self.cell_search()
                

    def add_price(self, count):
        pass#expenses[f'{celendar[day]}{count}'] = dict_product[product][0]

    def cell_search(self):
        count = 2
        while expenses[f'B{count}'].value is not None:
            count += 1
        self.add_price(self, count)


ch = Check(dict_product)
#ch.adding_elements(10)
print(ch)



        

wb.save('test.xlsx')
wb.close()
print('Good')







#'C:\\Users\\Alex_job\\Documents\\Git\\Test_site\\My.json'

#'C:\\Users\\Alex_job\\Documents\\Git\\Test_site\\programma_dly_obrabotki_chekov\\extract.json'