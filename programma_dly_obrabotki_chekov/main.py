import openpyxl
import json

#===============================================
#  Json файла с чеками

with open('C:\\Users\\Alex_job\\Documents\\Git\\Test_site\\programma_dly_obrabotki_chekov\\extract.json', mode= 'r', encoding= 'UTF-8') as file:
    my_file = json.load(file)
    new_json = json.dumps(my_file, ensure_ascii=False, indent= 4)
    my_json = json.loads(new_json)  

#===============================================
#  Кадры для созданрия файла exel и его заполнения (шаблон)

#wb = Workbook()
#income = wb.active
#expenses = wb.create_sheet('Расходы')
#income.title = 'Доходы'
#income['A1'] = product_name
#expenses['A1'] = 'Январь'
#expenses['A2'] = 'Категория'
#expenses['B2'] = 'Подкатегория'
#expenses['A3'] = 'Продукты'
celendar_day = {1:'c', 2:'d', 3:'e', 4:'f', 5:'g', 6:'h', 7:'i', 8:'j', 9:'k', 10:'l', 11:'m', 12:'n',
             13:'o', 14:'p', 15:'q', 16:'r', 17:'s', 18:'t', 19:'u', 20:'v', 21:'w', 22:'x', 23:'y',
             24:'z', 25:'aa', 26:'ab', 27:'ac', 28:'ad', 29:'ae', 30:'af', 31:'ag'}
celendar_month = {1:'Январь', 2:'Февраль', 3:'Март', 4:'Апрель', 5:'Май', 6:'Июнь',
                    7:'Июль', 8:'Август', 9:'Сентябрь', 10:'Октябрь', 11:'Ноябрь', 12:'Декабрь'}

#===============================================
#  Открытие сущетвующего файла exel и его наполнение чеками

wb = openpyxl.load_workbook(filename= 'C:\\Users\\Alex_job\\Documents\\Git\\Test_site\\programma_dly_obrabotki_chekov\\test.xlsx')

expenses = wb['Расходы']
income = wb['Доходы']


class Check:
    
    def __init__(self, check):
        self.dict_product = check
        self.scaning_products(self)

    def scaning_products(self, element):
        month_number = int(dict_product[list(dict_product.keys())[0]][2][5:7])
        day_number = int(dict_product[list(dict_product.keys())[0]][2][8:])
        month_line = self.search_month_number(celendar_month[month_number])
        for key in self.dict_product:
            line = self.product_search(key, month_line)
            if expenses[f'B{line}'].value == key:
                expenses[f'{celendar_day[day_number]}{line}'] = dict_product[key][0]
            else:
                expenses[f'B{line}'] = key
                expenses[f'{celendar_day[day_number]}{line}'] = dict_product[key][0]
                expenses.insert_rows(line + 1, 1)

    def search_month_number(self, month):
        count = 1
        while expenses[f'A{count}'].value != month:
            count += 1
        return count + 1
    
    def product_search(self, key, month):
        count = month
        while expenses[f'B{count}'].value != key and expenses[f'B{count}'].value is not None:
            count += 1
        return count

for i in range(len(my_json)):
    cheque = my_json[i]
    dict_product = {}
    for i in range(len(cheque["ticket"]["document"]["receipt"]["items"])):
        dict_product[cheque["ticket"]["document"]["receipt"]["items"][i]["name"]] = [cheque["ticket"]["document"]["receipt"]["items"][i]["price"] / 100,
                                                                                      cheque["ticket"]["document"]["receipt"]["items"][i]["quantity"],
                                                                                        cheque['createdAt'][0:-15]]
    ch = Check(dict_product)
  

wb.save('test.xlsx')
wb.close()
print('Good')







#'C:\\Users\\Alex_job\\Documents\\Git\\Test_site\\My.json'

#'C:\\Users\\Alex_job\\Documents\\Git\\Test_site\\programma_dly_obrabotki_chekov\\extract.json'